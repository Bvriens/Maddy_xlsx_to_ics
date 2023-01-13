import openpyxl
import icalendar
import datetime
from handlers.eventHandler import EventHandler

# Set up logging
import logging

logging.basicConfig(level=logging.DEBUG)


event_handler = EventHandler()


# function to convert excel date to python datetime
def excel_to_python_datetime(excel_date):
    # return datetime.datetime(1899, 12, 30) + datetime.timedelta(days=excel_date)
    return excel_date


morning_start_time = datetime.time(8, 35)
morning_end_time = datetime.time(12, 25)
afternoon_start_time = datetime.time(13, 5)
afternoon_end_time = datetime.time(16, 55)

# read xlsx file
workbook = openpyxl.load_workbook("IVA_2022.xlsx")
worksheet = workbook.active

# create ics calendar
calendar = icalendar.Calendar()
calendar.add("prodid", "-//My calendar product//mxm.dk//")
calendar.add("version", "2.0")

# trying to get the distinct course names from the classes
# list_of_classes = []
# # loop through rows and get al the classes strings
# for row in worksheet.iter_rows(values_only=True):
#     if row[0]:
#         pass
#     else:
#         print('no date pass')
#         continue
#     morning_class = row[1]
#     afternoon_class = row[2]
#     if morning_class:
#        list_of_classes.append(morning_class)

#     if afternoon_class:
#         list_of_classes.append(afternoon_class)

# # find the courses (maddy gebruikt niet exect dezelfde namen in de excel)
# print("classes", list_of_classes)
# list_of_courses = get_common_parts(list_of_classes)

# print("courses", list_of_courses)

# loop through rows and add events to calendar
for row in worksheet.iter_rows(values_only=True):
    if row[0]:
        date = excel_to_python_datetime(row[0]).date()
    else:
        print("no date pass")
        continue

    morning_course = row[1]
    afternoon_course = row[2]

    if morning_course:
        print("morning", morning_course, date, morning_start_time, morning_end_time)
        event_handler.addEvent(morning_course, date, morning_start_time, morning_end_time)

    if afternoon_course:
        print("afternoon", afternoon_course, date, afternoon_start_time, afternoon_end_time)
        event_handler.addEvent(afternoon_course, date, afternoon_start_time, afternoon_end_time)

logging.debug("*** printEvents *******************")
logging.debug(event_handler.printEvents)
logging.debug("***********************************")

# print all the courses and ask to select the courses
course_list = event_handler.getCourses()

for i, course in enumerate(course_list, 1):
    print(f"{i}. {course}")
selected_courses = input("Enter the numbers of the courses you want to save, separated by commas: ").split(",")
selected_courses = [int(course.strip()) for course in selected_courses]
selected_courses = [course_list[i - 1] for i in selected_courses]

# Get events according to selected courses
events_to_save = event_handler.getEventsList(selected_courses)

# create new calendar with only the selected events
calendar = icalendar.Calendar()
calendar.add("prodid", "-//My calendar product//mxm.dk//")
calendar.add("version", "2.0")
for event in events_to_save:
    calendar.add_component(event)

# write ics file
with open("courses.ics", "wb") as f:
    f.write(calendar.to_ical())

print(f"Successfully written {len(events_to_save)} events to courses.ics")
